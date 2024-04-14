import kivy
import os
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from faker import Faker
from openpyxl import Workbook
from openpyxl import load_workbook
import random

class PhoneGeneratorApp(App):
    def build(self):
        self.fake = Faker('pt_BR')
        self.layout = BoxLayout(orientation='vertical', spacing=10)
        
        self.label_prefixo = Label(text='Digite os 7 primeiros dígitos do telefone:')
        self.layout.add_widget(self.label_prefixo)

        self.prefixo_input = TextInput(multiline=False)
        self.layout.add_widget(self.prefixo_input)

        self.label_qtd = Label(text='Quantidade de números:')
        self.layout.add_widget(self.label_qtd)

        self.qtd_input = TextInput(multiline=False)
        self.layout.add_widget(self.qtd_input)

        self.gerar_button = Button(text='Gerar Telefones e Salvar', size_hint_y=None, height=40)
        self.gerar_button.bind(on_press=self.gerar_telefones_e_salvar)
        self.layout.add_widget(self.gerar_button)

        self.result_label = Label(text='')
        self.layout.add_widget(self.result_label)

        return self.layout

    def gerar_telefones_e_salvar(self, instance):
        prefixo = self.prefixo_input.text.strip()  # Pegando os 7 primeiros dígitos
        qtd = self.qtd_input.text.strip()
        
        if prefixo.isdigit() and len(prefixo) == 7 and qtd.isdigit():
            qtd = int(qtd)
            if 0 < qtd <= 500000:
                telefones = []
                for _ in range(qtd):
                    nome = self.fake.name()  # Nome diferente para cada número
                    ultimos_quatro = self.gerar_ultimos_quatro()
                    telefone = f'{prefixo}{ultimos_quatro}'  # Concatenando os 7 primeiros com os 4 últimos
                    telefones.append((nome, telefone))  # Armazenar nome e telefone em tupla
                self.result_label.text = f'Total de Telefones Gerados: {qtd}'
                self.salvar_no_excel(telefones)
            else:
                self.result_label.text = 'A quantidade deve ser um número entre 1 e 500000'
        else:
            self.result_label.text = 'Digite um prefixo válido (7 dígitos numéricos) e uma quantidade válida'

    def salvar_no_excel(self, telefones):
        filename = 'numeros_telefone.xlsx'
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.append(['Nome', 'Telefone'])
        else:
            wb = load_workbook(filename)
            ws = wb.active
        for nome, telefone in telefones:
            ws.append([nome, telefone])
        wb.save(filename)

    def gerar_ultimos_quatro(self):
        return ''.join(str(random.randint(0, 9)) for _ in range(4))

if __name__ == '__main__':
    PhoneGeneratorApp().run()
